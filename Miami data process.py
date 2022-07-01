# -*- coding: utf-8 -*-
"""
Created on Thu Jun 23 16:40:37 2022

@author: zhang
"""

from openpyxl import Workbook
from lxml import etree
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as stats
wb = Workbook()
ws1 = wb.create_sheet("MM",index=0)
#%% Building information (ID, building type, DEM, DSM, building height, footprint area, floor area, distance to coastline, population density )
df_1=pd.read_excel('Floor space calculation.xlsx',sheet_name='Miami',usecols='A,B,C,D,E,F,G')
Building_MM = df_1[df_1['FLOOR AREA (m2)']>0]
# Building_MM_D = Building_MM[(Building_MM['BUILDING TYPE'] == 'apartments')|(Building_MM['BUILDING TYPE'] == 'house')|(Building_MM['BUILDING TYPE'] == 'residential')|(Building_MM['BUILDING TYPE'] == 'yes')]
#GFA_D = Building_MM_D['FLOOR AREA (m2)'].sum()
df_2 =pd.read_excel('Distance from building to coastline.xlsx',sheet_name='MM') # distance from buildings to coastline (m)
df_3=pd.merge(Building_MM,df_2,how='inner',on=["ID"])
Population_density = pd.read_excel('Population density.xlsx',sheet_name='MM') # population density (capita/km2)
df_4 = pd.merge(df_3,Population_density,how='inner', on=["ID"])
df_4 =df_4[df_4['BUILDING HEIGHT (m)']>3.6]

#%% Determine the design cooling load based on prototype simulations (find the maximum requirement)
file_path='C:\EnergyPlusV9-6-0\Residential building\Residential buildings results\Miami\Monthly load\\' 
all_excel=os.listdir(file_path) 
cooling_load = np.zeros([12,1])
col = 0
for excel_num in all_excel:
            col +=1
            data = pd.read_excel(file_path+excel_num, sheet_name='Sheet1', na_values='n/a')
            data.drop(data.index[13:], inplace=True)
            data.dropna(axis=0, how='any', thresh=None, subset=None, inplace=False)
            a = data.iloc[1:].astype(float)
            b = a.sum(axis=1)/1000 # total cooling load of all zones (kW)
            b = np.expand_dims(b, axis=1)
            cooling_load = np.append(cooling_load, b, axis=1)
cooling_load = (cooling_load/7059) # kW/m2 (residential)
c= cooling_load.mean(axis=1)
k= cooling_load.std(axis=1)
design_cooling_capacity= c.max() # annual maximum requirement (kW/m2) (residential)

annual_energy_saving_DCS = (df_4['FLOOR AREA (m2)']*design_cooling_capacity*1.1*24*365/2.7-df_4['FLOOR AREA (m2)']*design_cooling_capacity*1.1*24*365/5)
annual_total_DCS = annual_energy_saving_DCS.sum()
ws1.cell(row=1, column=1, value=annual_total_DCS)

#%% DCS costs of each building
Discount_rate = 0.005  #senstivity parameter +-10%
lifespan = 60 #senstivity parameter +-10%
ele_price = 0.159 #senstivity parameter +-10%

Cooling_load_required =  df_4['FLOOR AREA (m2)']*design_cooling_capacity # cooling load required of each building (kW)
f = 14985 # based on the DCS costs in HK
cost_each_building_DCS = (f* (Cooling_load_required**0.87))*1.3*0.93 # 0.87 is the scaling factor 
lump_sum = annual_energy_saving_DCS*ele_price-cost_each_building_DCS*0.013
PVaule_saving = lump_sum*((Discount_rate+1)**lifespan-1)/(Discount_rate*(1+Discount_rate)**lifespan) # present-value of 70-year electricity savings

Net_cost_DCS = cost_each_building_DCS-PVaule_saving
e = pd.DataFrame(np.expand_dims(Net_cost_DCS, axis=1),columns=['DCS costs (USD)'])
g = e.drop(0,axis=0)
df_5 = (df_4).join(g)
#np.savetxt('C:\EnergyPlusV9-6-0\Residential building\Residential buildings results\HK_all_data.csv', df_5, fmt="%s", delimiter=" ")
cost_total_DCS = e.sum()
 #kWh/a
ws1.cell(row=1, column=2, value=cost_total_DCS[0])
#%% PV solar panel cost of each building
lifespan_PV = 25
suitable_ratio = 0.5 #adjustable
annual_AC_output = 6296 #kWh per 7kW panel (50 m2 each)
annual_ele_gain_PV = (df_5['FOOTPRINT AREA (m2)']*suitable_ratio)/110*annual_AC_output


cost_each_building_PV = 312*(df_5['FOOTPRINT AREA (m2)']*suitable_ratio)**0.9
lump_sum_PV = annual_ele_gain_PV*ele_price-cost_each_building_PV*0.00454
PVaule_saving_PV = lump_sum_PV*((Discount_rate+1)**lifespan_PV-1)/(Discount_rate*(1+Discount_rate)**lifespan_PV)
Net_cost_PV = cost_each_building_PV - PVaule_saving_PV
annual_total_PV = annual_ele_gain_PV.sum()
h = pd.DataFrame(np.expand_dims(Net_cost_PV, axis=1),columns=['PV solar panel costs (USD)'])
df_6 = (df_5).join(h)
cost_total_PV = Net_cost_PV.sum()
ws1.cell(row=2, column=1, value=annual_total_PV)
ws1.cell(row=2, column=2, value=cost_total_PV)
#%% DCS GHG profile 
specific_HC = 4.184 # kJ/(kg*C)
delt_T = 5
density = 1000 # kg/m3
gg = 9.81
elevation = df_6['DSM (s)']
pum_eff = 0.6
GHG_factor = 0.423 # kg CO2e/kWh adjustable
chill_water_flow_rate = df_4['FLOOR AREA (m2)']*design_cooling_capacity*24*365*3600/specific_HC/delt_T/density/365/24  # m3/h
pumping_power_chill = (chill_water_flow_rate*density*gg*elevation/(3.6*10**6))/pum_eff  # kW
cooling_water_flow_rate = df_4['FLOOR AREA (m2)']*design_cooling_capacity*24*365*3600/0.8/specific_HC/delt_T/density/365/24 # m3/h
pumping_power_cool = (cooling_water_flow_rate*density*gg*20/(3.6*10**6))/pum_eff  # kW
annual_pump_energy = (pumping_power_chill+pumping_power_cool)*24*365 #kWh/a
annual_GHG_reduction_DCS = (annual_energy_saving_DCS-annual_pump_energy)*GHG_factor*60 #kg CO2e
annual_total_GHG_red_DCS = annual_GHG_reduction_DCS.sum()
annual_GHG_reduction_DCS.name = 'annual_GHG_reduction_DCS'
df_7 = df_6.join(annual_GHG_reduction_DCS)
ws1.cell(row=1, column=3, value=annual_total_GHG_red_DCS)


#%% PV solar panel GHG profile
annual_GHG_reduction_PV = annual_ele_gain_PV*GHG_factor*25 # kg CO2e
annual_total_GHG_red_PV = annual_GHG_reduction_PV.sum()
annual_GHG_reduction_PV.name = 'annual_GHG_reduction_PV'
df_8 = df_7.join(annual_GHG_reduction_PV)
ws1.cell(row=2, column=3, value=annual_total_GHG_red_PV)      
# wb.save('DCS_PV_comparison_1.xlsx')


#%% eco-efficiency calculation
ee_DCS = Net_cost_DCS/annual_GHG_reduction_DCS # USD/kgCO2
ee_PV = Net_cost_PV/annual_GHG_reduction_PV
ee_DCS = pd.DataFrame(ee_DCS,columns=['ee_DCS (USD/kgCO2e)'])
ee_PV = pd.DataFrame(np.expand_dims(ee_PV, axis=1),columns=['ee_PV (USD/kgCO2e)'])
Pop_den = df_6['Population density']
Costline_dis = df_6['NEAR_DIST']
df_9 =df_8.join(ee_DCS)
df_10 =df_9.join(pd.DataFrame(ee_PV))
#np.savetxt('C:\EnergyPlusV9-6-0\Residential building\Residential buildings results\HK_data.csv', df_8, fmt="%s", delimiter=" ")
# df_10.to_excel('Miami data.xlsx')

#%% 0 as cutoff
# df_11 = df_10[(df_10['ee_DCS (USD/kgCO2e)']<0)]
# df_12 = df_10[(df_10['ee_PV (USD/kgCO2e)']<0)]
# ee_new = (df_11['DCS costs (USD)'].sum()+df_12['PV solar panel costs (USD)'].sum())/(df_11['annual_GHG_reduction_DCS'].sum()+df_12['annual_GHG_reduction_PV'].sum())


c_DCS = (df_10['DCS costs (USD)'].sum())/(df_10['annual_GHG_reduction_DCS'].sum())
c_PV = df_10['PV solar panel costs (USD)'].sum()/df_10['annual_GHG_reduction_PV'].sum()
#%% correlation analysis
# data_1 = df_10['Population density']
# data_2 = df_10['ee_DCS (USD/kgCO2e)']
# # data_1.corr(data_2)
# stats.spearmanr(data_1,data_2)
